# -*- encoding: utf-8 -*-
'''
@File    :   sheet_copy.py
@Author  :   北极星光 
@Contact :   light22@126.com
'''

from excel_operate import ExcelOperate
from copy import copy
from openpyxl.utils import get_column_letter


class SheetCopy:
    def __init__(self, src_file_path: str = None, tag_file_path: str = None, sheet_name: str = None, column_adjust: float = 0) -> None:
        '''===================================\n
        src_file_path: 此参数为源Excel文件路径
        tag_file_path: 此参数为目标Excel文件路径，不指定的情况下默认新建工作薄
        sheet_name: 此参数为被复制的工作表名称，不指定的情况下默认为源工作表中的当前激活工作表
        column_adjust: 此参数为列宽修正系数，作用为修正列宽误差
        '''
        self.src_file = ExcelOperate(src_file_path, sheet_name)
        if tag_file_path == None:
            self.tag_file = ExcelOperate(sheet_name=self.src_file.ws.title)
        else:
            self.tag_file = ExcelOperate(tag_file_path)
            if self.src_file.ws.title not in self.tag_file.wb.sheetnames:
                self.tag_file.wb.create_sheet(self.src_file.ws.title)
                self.tag_file.ws = self.tag_file.wb[self.src_file.ws.title]
            else:
                n = 1
                while True:
                    if f'{self.src_file.ws.title} ({n})' not in self.tag_file.wb.sheetnames:
                        self.tag_file.wb.create_sheet(
                            f'{self.src_file.ws.title} ({n})')
                        self.tag_file.ws = self.tag_file.wb[
                            f'{self.src_file.ws.title} ({n})']
                        break
                    n += 1
        # 列宽误差修正系数，默认为0，由于openpyxl在设置列宽时与实际存在出入，因此可以给定此参数来调整误差
        self.column_adjust = column_adjust

    # 定义工作表复制模块
    def copy_sheet(self, src_file: ExcelOperate = None, tag_file: ExcelOperate = None, origin_row: int = 1, origin_col: int = 1):
        '''===================================\n
        复制工作表
        src_file: 原文件的ExcelOperate对象
        tag_file: 目标文件的ExcelOperate对象
        origin_row: 设定目标工作表写入的起始行，默认从第1行开始写入，传入大于或等于1的整数则会从设定的行开始写入。
        origin_col: 设定目标工作表写入的起始列，默认从第1列开始写入，传入大于或等于1的整数则会从设定的列开始写入。
        '''
        if src_file == None:
            src_file = self.src_file
        if tag_file == None:
            tag_file = self.tag_file
        for row in src_file.ws:
            # 遍历源xlsx文件制定sheet中的所有单元格
            for cell in row:  # 复制数据
                cell.row += origin_row - 1
                cell.column += origin_col - 1
                tag_file.ws[cell.coordinate].value = cell.value
                if cell.has_style:  # 复制样式
                    tag_file.ws[cell.coordinate].font = copy(cell.font)
                    tag_file.ws[cell.coordinate].border = copy(cell.border)
                    tag_file.ws[cell.coordinate].fill = copy(cell.fill)
                    tag_file.ws[cell.coordinate].number_format = copy(
                        cell.number_format)
                    tag_file.ws[cell.coordinate].protection = copy(
                        cell.protection)
                    tag_file.ws[cell.coordinate].alignment = copy(
                        cell.alignment)

        merged_cells = src_file.ws.merged_cells.ranges  # 已合并的单元格列表
        if len(merged_cells) > 0:  # 检测源xlsx中合并的单元格
            for merged_cell in merged_cells:
                tag_file.ws.merge_cells(start_row=merged_cell.min_row + origin_row - 1, end_row=merged_cell.max_row + origin_row - 1,
                                        start_column=merged_cell.min_col + origin_col - 1, end_column=merged_cell.max_col + origin_col - 1)  # 合并单元格
        # 开始处理行高列宽
        for i in range(1, src_file.ws.max_row + 1):
            tag_file.ws.row_dimensions[i + origin_row -
                                       1].height = src_file.ws.row_dimensions[i].height
        for i in range(1, src_file.ws.max_column + 1):
            tag_file.ws.column_dimensions[get_column_letter(
                i + origin_col - 1)].width = src_file.ws.column_dimensions[get_column_letter(i)].width + self.column_adjust  # 修正列宽误差

        return tag_file


# 调试
if __name__ == '__main__':
    src_file_path = r'D:\codes\Python Projects\Excel-operate\tests\示例.xlsx'
    copyer = SheetCopy(src_file_path)
    tag_file = copyer.copy_sheet(origin_col=5, origin_row=8)
    tag_file.save(r'D:\Desktop\1111.xlsx')
