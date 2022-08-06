from openpyxl import Workbook,load_workbook
from copy import copy
from openpyxl.utils import get_column_letter

class ExcelOperate:
    def __init__(self,file_path:str=None,sheet_name:str=None):
        '''
        file_path: 指定目标Excel文件路径，不指定的话则默认打开新建工作薄
        sheet_name: 指定目标工作表名称，不指定的话默认为当前激活工作表，如果参数file_path未指定的话，请不要给本参数赋值
        '''
        self.file_path = file_path
        if file_path == None:
            self.wb = Workbook()
        else:
            self.wb = load_workbook(file_path)
        
        if sheet_name == None:
            self.ws = self.wb.active
        else:
            self.ws = self.wb[sheet_name]

    def insert_rows(self,idx:int,amount:int=1,height:float='before'): # 定义插入行模块
        '''
        idx: 参数idx为指定插入行的起始位置
        amount: 参数amount为指定插入的行数，不指定的情况下默认插入1行
        height: 参数height为指定插入行的行高，不指定的情况下默认继承前一行的行高，指定数字可以设置对应的行高，指定None为自动行高
        '''
        row_height_list = [] # 定义空列表用以接收从插入行开始到最大行的所有行高数据
        for i in range(idx,self.ws.max_row+1):
            row_height_list.append(self.ws.row_dimensions[i].height)

        merged_cells = self.ws.merged_cell_ranges # 当前工作表中所有已合并的单元格列表
        if len(merged_cells) > 0: # 判断是否存在合并单元格
            merged_cell_list = [] # 定义空列表，方便储存插入行后的已合并单元格信息
            for merged_cell in merged_cells:
                merged_cell_info = {}
                if merged_cell.min_row >= idx:
                    merged_cell_info['min_row'] = merged_cell.min_row + amount
                else:
                    merged_cell_info['min_row'] = merged_cell.min_row
                
                if merged_cell.max_row >= idx:
                    merged_cell_info['max_row'] = merged_cell.max_row + amount
                else:
                    merged_cell_info['max_row'] = merged_cell.max_row

                merged_cell_info['min_col'] = merged_cell.min_col
                merged_cell_info['max_col'] = merged_cell.max_col

                merged_cell_list.append(merged_cell_info) # 将插入行后的合并单元格信息存入列表

                self.ws.unmerge_cells(start_row=merged_cell.min_row,end_row=merged_cell.max_row,start_column=merged_cell.min_col,end_column=merged_cell.max_col) # 解除合并单元格

            self.ws.insert_rows(idx,amount) # 插入空白行

            for i in merged_cell_list:
                self.ws.merge_cells(start_row=i['min_row'],end_row=i['max_row'],start_column=i['min_col'],end_column=i['max_col']) # 重新合并单元格
                
        else:
            self.ws.insert_rows(idx,amount) # 插入空白行

        for src_row in self.ws.iter_rows(min_row=idx-1,max_row=idx-1): #从插入行的上方复制单元格格式
            for tag_row in self.ws.iter_rows(min_row=idx,max_row=idx+amount-1):
                for src_cell,tag_cell in zip(src_row,tag_row):
                    tag_cell.font = copy(src_cell.font) # 复制字体
                    tag_cell.border = copy(src_cell.border) # 复制边框
                    tag_cell.fill = copy(src_cell.fill) # 复制填充色
                    tag_cell.number_format = copy(src_cell.number_format) # 复制数据格式
                    tag_cell.protection = copy(src_cell.protection) # 复制保护状态
                    tag_cell.alignment = copy(src_cell.alignment) # 复制对齐格式
        
        for i in range(idx,idx+amount+1): # 设置插入行的行高
            if height == 'before':
                self.ws.row_dimensions[i].height = self.ws.row_dimensions[idx-1].height
            else:
                self.ws.row_dimensions[i].height = height

        for i,h in zip(range(idx+amount,self.ws.max_row+1),row_height_list):
            self.ws.row_dimensions[i].height = h

    def insert_cols(self,idx:int,amount:int=1,width:float='before'): # 定义插入列模块
        '''
        idx: 参数idx为指定插入列起始的位置
        amount: 参数amount为指定插入的列数，不指定的情况下默认插入1列
        width: 参数width为指定插入列的列宽，不指定的情况下默认继承前一列的列宽，指定数字可以设置对应的列宽
        '''
        col_width_list = [] # 定义空列表用以接收从插入列开始到最大列的所有列宽数据
        for i in range(idx,self.ws.max_column+1):
            col_width_list.append(self.ws.column_dimensions[get_column_letter(i)].width)
        
        merged_cells = self.ws.merged_cell_ranges # 当前工作表中所有已合并的单元格列表
        if len(merged_cells) > 0: # 判断是否存在合并单元格
            merged_cell_list = [] # 定义空列表，方便储存插入列后的已合并单元格信息
            for merged_cell in merged_cells:
                merged_cell_info = {}
                if merged_cell.min_col >= idx:
                    merged_cell_info['min_col'] = merged_cell.min_col + amount
                else:
                    merged_cell_info['min_col'] = merged_cell.min_col
                
                if merged_cell.max_col >= idx:
                    merged_cell_info['max_col'] = merged_cell.max_col + amount
                else:
                    merged_cell_info['max_col'] = merged_cell.max_col

                merged_cell_info['min_row'] = merged_cell.min_row
                merged_cell_info['max_row'] = merged_cell.max_row

                merged_cell_list.append(merged_cell_info) # 将插入列后的合并单元格信息存入列表

                self.ws.unmerge_cells(start_row=merged_cell.min_row,end_row=merged_cell.max_row,start_column=merged_cell.min_col,end_column=merged_cell.max_col) # 解除合并单元格

            self.ws.insert_cols(idx,amount) # 插入空白列

            for i in merged_cell_list:
                self.ws.merge_cells(start_row=i['min_row'],end_row=i['max_row'],start_column=i['min_col'],end_column=i['max_col']) # 重新合并单元格
                
        else:
            self.ws.insert_cols(idx,amount) # 插入空白列

        for src_col in self.ws.iter_cols(min_col=idx-1,max_col=idx-1): #从插入列的前方复制单元格格式
            for tag_col in self.ws.iter_cols(min_col=idx,max_col=idx+amount-1):
                for src_cell,tag_cell in zip(src_col,tag_col):
                    tag_cell.font = copy(src_cell.font) # 复制字体
                    tag_cell.border = copy(src_cell.border) # 复制边框
                    tag_cell.fill = copy(src_cell.fill) # 复制填充色
                    tag_cell.number_format = copy(src_cell.number_format) # 复制数据格式
                    tag_cell.protection = copy(src_cell.protection) # 复制保护状态
                    tag_cell.alignment = copy(src_cell.alignment) # 复制对齐格式

        for i in range(idx,idx+amount+1): # 设置插入列的列宽
            if width == 'before':
                self.ws.column_dimensions[get_column_letter(i)].width = self.ws.column_dimensions[get_column_letter(idx-1)].width
            else:
                self.ws.column_dimensions[get_column_letter(i)].width = width

        for i,w in zip(range(idx+amount,self.ws.max_column+1),col_width_list):
            self.ws.column_dimensions[get_column_letter(i)].width = w

    def delete_rows(self,idx:int,amount:int=1): # 定义删除行模块
        '''
        idx: 参数idx为指定删除行的起始位置
        amount: 参数amount为指定删除的行数，不指定的情况下默认删除1行
        '''
        row_height_list = [] # 定义空列表用以接收从删除行结尾开始到最大行的所有行高数据
        for i in range(idx+amount,self.ws.max_row+1):
            row_height_list.append(self.ws.row_dimensions[i].height)

        merged_cells = self.ws.merged_cell_ranges # 当前工作表中所有已合并的单元格列表
        if len(merged_cells) > 0: # 判断是否存在合并单元格
            merged_cell_list = [] # 定义空列表，方便储存删除行后的已合并单元格信息
            for merged_cell in merged_cells:
                merged_cell_info = {}
                if merged_cell.min_row >= idx and merged_cell.max_row <= idx + amount - 1:
                    self.ws.unmerge_cells(start_row=merged_cell.min_row,end_row=merged_cell.max_row,start_column=merged_cell.min_col,end_column=merged_cell.max_col) # 解除合并单元格
                    continue
                if merged_cell.min_row >= idx:
                    if merged_cell.min_row - amount < idx:
                        merged_cell_info['min_row'] = idx
                    else:
                        merged_cell_info['min_row'] = merged_cell.min_row - amount
                else:
                    merged_cell_info['min_row'] = merged_cell.min_row
                
                if merged_cell.max_row >= idx:
                    if merged_cell.max_row - amount < idx - 1:
                        merged_cell_info['max_row'] = idx - 1
                    else:
                        merged_cell_info['max_row'] = merged_cell.max_row - amount
                else:
                    merged_cell_info['max_row'] = merged_cell.max_row

                merged_cell_info['min_col'] = merged_cell.min_col
                merged_cell_info['max_col'] = merged_cell.max_col

                merged_cell_list.append(merged_cell_info) # 将删除行后的合并单元格信息存入列表

                self.ws.unmerge_cells(start_row=merged_cell.min_row,end_row=merged_cell.max_row,start_column=merged_cell.min_col,end_column=merged_cell.max_col) # 解除合并单元格

            self.ws.delete_rows(idx,amount) # 删除行

            for i in merged_cell_list:
                self.ws.merge_cells(start_row=i['min_row'],end_row=i['max_row'],start_column=i['min_col'],end_column=i['max_col']) # 重新合并单元格
                
        else:
            self.ws.delete_rows(idx,amount) # 删除行

        for i,h in zip(range(idx,self.ws.max_row+1),row_height_list): # 恢复删除行后的行高
            self.ws.row_dimensions[i].height = h

    def delete_cols(self,idx:int,amount:int=1): # 定义删除列模块
        '''
        idx: 参数idx为指定删除列的起始位置
        amount: 参数amount为指定删除的列数，不指定的情况下默认删除1列
        '''
        col_width_list = [] # 定义空列表用以接收从删除列结尾开始到最大列的所有列宽数据
        for i in range(idx+amount,self.ws.max_column+1):
            col_width_list.append(self.ws.column_dimensions[get_column_letter(i)].width)
        
        merged_cells = self.ws.merged_cell_ranges # 当前工作表中所有已合并的单元格列表
        if len(merged_cells) > 0: # 判断是否存在合并单元格
            merged_cell_list = [] # 定义空列表，方便储存删除列后的已合并单元格信息
            for merged_cell in merged_cells:
                merged_cell_info = {}
                if merged_cell.min_col >= idx and merged_cell.max_col <= idx + amount - 1:
                    self.ws.unmerge_cells(start_row=merged_cell.min_row,end_row=merged_cell.max_row,start_column=merged_cell.min_col,end_column=merged_cell.max_col) # 解除合并单元格
                    continue
                if merged_cell.min_col >= idx:
                    if merged_cell.min_col - amount < idx:
                        merged_cell_info['min_col'] = idx
                    else:
                        merged_cell_info['min_col'] = merged_cell.min_col - amount
                else:
                    merged_cell_info['min_col'] = merged_cell.min_col
                
                if merged_cell.max_col >= idx:
                    if merged_cell.max_col - amount < idx - 1:
                        merged_cell_info['max_col'] = idx - 1
                    else:
                        merged_cell_info['max_col'] = merged_cell.max_col - amount
                else:
                    merged_cell_info['max_col'] = merged_cell.max_col

                merged_cell_info['min_row'] = merged_cell.min_row
                merged_cell_info['max_row'] = merged_cell.max_row

                merged_cell_list.append(merged_cell_info) # 将插入列后的合并单元格信息存入列表

                self.ws.unmerge_cells(start_row=merged_cell.min_row,end_row=merged_cell.max_row,start_column=merged_cell.min_col,end_column=merged_cell.max_col) # 解除合并单元格

            self.ws.delete_cols(idx,amount) # 删除列

            for i in merged_cell_list:
                self.ws.merge_cells(start_row=i['min_row'],end_row=i['max_row'],start_column=i['min_col'],end_column=i['max_col']) # 重新合并单元格
                
        else:
            self.ws.delete_cols(idx,amount) # 删除列

        for i,w in zip(range(idx,self.ws.max_column+1),col_width_list): # 恢复删除行后的列宽
            self.ws.column_dimensions[get_column_letter(i)].width = w


    def save(self,file_path:str=None):
        '''
        file_path: 指定文件保存路径，如果对象为已有工作薄的话，此参数不指定则默认为覆盖原文件保存，如果对象为新建工作薄的话此参数必须指定，否则会报错
        '''
        if file_path == None:
            self.wb.save(self.file_path)
        else:
            self.wb.save(file_path)

# 调试
if __name__ == '__main__':
    excel = ExcelOperate()

    excel.ws['C15'].value = '测试通过！！'
    excel.wb.save('d:/desktop/111.xlsx')