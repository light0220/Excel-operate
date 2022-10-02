# -*- encoding: utf-8 -*-
'''
@File    :   sheet_comparison.py
@Author  :   北极星光 
@Contact :   light22@126.com
'''

from excel_operate import ExcelOperate
from excel_operate import SheetCopy
from excel_operate.list_operate import *
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Side, Border, Font


class SheetComparison:
    def __init__(self, src_excel: ExcelOperate, cmp_excel: ExcelOperate) -> None:
        '''---
        ### 传入两个Excel对象，进行比较。生成报告
        ---
        + src_excel: 原Excel对象，可传入ExcelOperate类对象
        + cmp_excel: 待比较的Excel对象，可传入ExcelOperate类对象
        '''
        self.src_excel = src_excel
        self.cmp_excel = cmp_excel

    def set_title_row(self, src_title_row: int, cmp_title_row: int):
        '''---
        ### 此方法可以为原工作表及待比较工作表设置表头行
        ---
        + src_title_row: 原Excel工作表的表头行，可传入整数类型
        + cmp_title_row: 待比较Excel工作表的表头行，可传入整数类型
        '''
        self.src_title_row = src_title_row
        self.cmp_title_row = cmp_title_row

    def set_key_col(self, src_key_col: int, cmp_key_col: int):
        '''---
        ### 此方法可以为原工作表及待比较工作表设置关键列
        ---
        + src_key_col: 原Excel工作表的关键列，可传入整数类型
        + cmp_key_col: 待比较Excel工作表的关键列，可传入整数类型
        '''
        self.src_key_col = src_key_col
        self.cmp_key_col = cmp_key_col

    def font_color(self, cell, style):
        '''---
        ### 此方法为目标单元格设置样式
        ---
        + cell: 目标单元格，请传入单元格对象
        + style: 目标单元格将要设置成的样式，只支持'zeng','shan'和'gai'这三个参数。
        '''
        if style == 'zeng':
            cell.font = Font(color='0000FF')
        if style == 'shan':
            cell.font = Font(color='FF0000', strike=True)
        if style == 'gai':
            cell.font = Font(color='FF00FF')

    def get_title_list(self):
        src_title_list = [
            i.value for i in self.src_excel.ws[self.src_title_row]]
        cmp_title_list = [
            i.value for i in self.cmp_excel.ws[self.cmp_title_row]]
        src_title_list = list_replace(src_title_list, None, '空')
        cmp_title_list = list_replace(cmp_title_list, None, '空')
        return src_title_list, cmp_title_list

    def get_key_list(self):
        src_key_list = [i.value for idx, i in enumerate(
            self.src_excel.ws[get_column_letter(self.src_key_col)], 1) if idx > self.src_title_row]
        cmp_key_list = [i.value for idx, i in enumerate(
            self.cmp_excel.ws[get_column_letter(self.cmp_key_col)], 1) if idx > self.cmp_title_row]
        src_key_list = list_replace(src_key_list, None, '空')
        cmp_key_list = list_replace(cmp_key_list, None, '空')
        return src_key_list, cmp_key_list

    def compare(self):
        '''---
        ### 对比工作表：将原工作表及目标工作表的表头行和关键列设置好之后即可使用此方法对比工作表，并返回对比报告。
        ---
        '''
        # 如果目标工作表的表头行与原工作表的表头行不在同一行
        if self.src_title_row != self.cmp_title_row:
            if self.src_title_row > self.cmp_title_row:
                self.cmp_excel.insert_rows(
                    1, self.src_title_row-self.cmp_title_row)
                self.cmp_title_row = self.src_title_row
            else:
                self.src_excel.insert_rows(
                    1, self.cmp_title_row-self.src_title_row)
                self.src_title_row = self.cmp_title_row

        # 如果目标工作表的关键列与原工作表的关键列不在同一列
        if self.src_key_col != self.cmp_key_col:
            if self.src_key_col > self.cmp_key_col:
                self.cmp_excel.insert_rows(
                    1, self.src_key_col-self.cmp_key_col)
                self.cmp_key_col = self.src_key_col
            else:
                self.src_excel.insert_rows(
                    1, self.cmp_key_col-self.src_key_col)
                self.src_key_col = self.cmp_key_col

        src_title_list, cmp_title_list = self.get_title_list()

        # 如果目标工作表的表头行与原工作表的表头行不相同
        if src_title_list != cmp_title_list:
            self.src_excel.insert_rows(1)
            self.src_title_row += 1
            self.cmp_excel.insert_rows(1)
            self.cmp_title_row += 1
            src_title_list = duplicate_to_only(
                src_title_list)  # 通过重命名的方式给列表去重。
            cmp_title_list = duplicate_to_only(
                cmp_title_list)  # 通过重命名的方式给列表去重。
            src_title_list, cmp_title_list = list_matching(
                src_title_list, cmp_title_list)  # 通过插入占位元素，将两列表相匹配。

            for i in range(len(src_title_list)):
                if src_title_list[i] == None:
                    self.src_excel.insert_cols(i + 1)
            for i in range(len(cmp_title_list)):
                if cmp_title_list[i] == None:
                    self.cmp_excel.insert_cols(i + 1)
            for idx, i, j in zip(range(len(src_title_list)), src_title_list, cmp_title_list):
                if i == None:
                    self.src_excel.ws[1][idx].value = '临'
                    self.cmp_excel.ws[1][idx].value = '增'
                if j == None:
                    self.src_excel.ws[1][idx].value = '删'
                    self.cmp_excel.ws[1][idx].value = '临'

            # 设置对齐、边框及字体颜色
            for cell in self.src_excel.ws[1]:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.border = Border(left=Side('thin'), right=Side(
                    'thin'), top=Side('thin'), bottom=Side('thin'))
            for cell in self.cmp_excel.ws[1]:
                cell.alignment = Alignment(
                    horizontal='center', vertical='center')
                cell.border = Border(left=Side('thin'), right=Side(
                    'thin'), top=Side('thin'), bottom=Side('thin'))
            for col in self.src_excel.ws.iter_cols():
                if col[0].value == '删':
                    for cell in col:
                        self.font_color(cell, 'shan')
            for col in self.cmp_excel.ws.iter_cols():
                if col[0].value == '增':
                    for cell in col:
                        self.font_color(cell, 'zeng')

        src_key_list, cmp_key_list = self.get_key_list()
        # 如果目标工作表的关键列与原工作表的关键列不相同
        if src_key_list != cmp_key_list:
            src_key_list = duplicate_to_only(src_key_list)  # 通过重命名的方式给列表去重。
            cmp_key_list = duplicate_to_only(cmp_key_list)  # 通过重命名的方式给列表去重。
            src_key_list, cmp_key_list = list_matching(
                src_key_list, cmp_key_list)  # 通过插入占位元素，将两列表相匹配。

            for i in range(len(src_key_list)):
                if src_key_list[i] == None:
                    self.src_excel.insert_rows(i + self.src_title_row + 1)
            for i in range(len(cmp_key_list)):
                if cmp_key_list[i] == None:
                    self.cmp_excel.insert_rows(i + self.cmp_title_row + 1)

        # 对比行，在原工作表行首写入标记同时设置字体颜色
        self.src_excel.insert_cols(1, width=3)  # 在原工作表首列处插入地行空列用于标记'增','删','改'
        for cell in self.src_excel.ws['A']:
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = Border(left=Side('thin'), right=Side(
                'thin'), top=Side('thin'), bottom=Side('thin'))
        self.src_key_col += 1
        src_key_list = [i.value for idx, i in enumerate(
            self.src_excel.ws[get_column_letter(self.src_key_col)], 1) if idx > self.src_title_row]
        cmp_key_list = [i.value for idx, i in enumerate(
            self.cmp_excel.ws[get_column_letter(self.cmp_key_col)], 1) if idx > self.cmp_title_row]

        for idx, i, j in zip(range(self.src_title_row + 1, len(src_key_list) + self.src_title_row + 1), src_key_list, cmp_key_list):
            if i == None != j:
                self.src_excel.ws[f'A{idx}'].value = '增'
            if i != None == j:
                self.src_excel.ws[f'A{idx}'].value = '删'
        # 逐项对比数据，标注改动的地方
        for x in range(1, self.src_excel.ws.max_column):
            for y in range(2, self.src_excel.ws.max_row + 1):
                if self.src_excel.ws[f'{get_column_letter(x + 1)}{1}'].value not in ['增', '删', '临'] and self.src_excel.ws[f'A{y}'].value not in ['增', '删', '临'] and self.cmp_excel.ws[f'{get_column_letter(x)}{1}'].value not in ['增', '删', '临']:
                    if self.src_excel.ws[f'{get_column_letter(x + 1)}{y}'].value != self.cmp_excel.ws[f'{get_column_letter(x)}{y}'].value:
                        self.src_excel.ws[f'A{y}'].value = '改'
                        self.font_color(self.src_excel.ws[f'A{y}'], 'gai')
                        if self.src_excel.ws[f'{get_column_letter(x + 1)}{y}'].value == None:
                            self.font_color(
                                self.cmp_excel.ws[f'{get_column_letter(x)}{y}'], 'zeng')
                        elif self.cmp_excel.ws[f'{get_column_letter(x)}{y}'].value == None:
                            self.font_color(
                                self.src_excel.ws[f'{get_column_letter(x + 1)}{y}'], 'shan')
                        else:
                            self.font_color(
                                self.src_excel.ws[f'{get_column_letter(x + 1)}{y}'], 'gai')
                            self.font_color(
                                self.cmp_excel.ws[f'{get_column_letter(x)}{y}'], 'gai')
        # 删除临时行
        for i in range(self.src_excel.ws.max_column, 0, -1):
            if self.src_excel.ws[f'{get_column_letter(i)}1'].value == '临':
                self.src_excel.delete_cols(i)
        for i in range(self.cmp_excel.ws.max_column, 0, -1):
            if self.cmp_excel.ws[f'{get_column_letter(i)}1'].value == '临':
                self.cmp_excel.delete_cols(i)

        # 将原工作表及目标工作表合并后写入报告文件
        self.src_excel.insert_rows(1, height=20)
        self.src_excel.ws['A1'].value = '原 工 作 表'
        self.src_excel.ws['A1'].font = Font(name='微软雅黑', size=16, b=True)
        self.src_excel.ws['A1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.src_excel.ws.merge_cells(
            f'A1:{get_column_letter(self.src_excel.ws.max_column)}1')
        self.cmp_excel.insert_rows(1, height=20)
        self.cmp_excel.ws['A1'].value = '目 标 工 作 表'
        self.cmp_excel.ws['A1'].font = Font(name='微软雅黑', size=16, b=True)
        self.cmp_excel.ws['A1'].alignment = Alignment(
            horizontal='center', vertical='center')
        self.cmp_excel.ws.merge_cells(
            f'A1:{get_column_letter(self.cmp_excel.ws.max_column)}1')

        rpt = ExcelOperate(sheet_name='对比结果')
        cp = SheetCopy()
        rpt = cp.copy_sheet(src_file=self.src_excel, tag_file=rpt)
        rpt = cp.copy_sheet(src_file=self.cmp_excel, tag_file=rpt,
                            origin_col=self.src_excel.ws.max_column + 1)
        for row in rpt.ws.iter_rows():
            if row[0].value == '增':
                for cell in row:
                    self.font_color(cell, 'zeng')
            if row[0].value == '删':
                for cell in row:
                    self.font_color(cell, 'shan')
        return rpt


# 调试
if __name__ == '__main__':
    src_excel = ExcelOperate('tests\比较示例 - 原.xlsx')
    cmp_excel = ExcelOperate('tests\比较示例 - 对比.xlsx')
    report_path = 'D:/Desktop/对比报告.xlsx'
    cmper = SheetComparison(src_excel, cmp_excel)
    cmper.set_title_row(2, 2)
    cmper.set_key_col(2, 2)
    report = cmper.compare()
    report.save(report_path)
    print('对比完成！对比报告已保存至：', report_path)
